from bot.db.requests import add_regular_lessons, add_uday_lessons
from bot.db.requests import delete_old_schedules, delete_repeated_schedule
from bot.db.requests import get_all_users, delete_user


from .lesson import Lesson

from aiogram.exceptions import TelegramForbiddenError
from aiogram import Bot

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from requests import Session, post
from typing import Union
from bs4 import BeautifulSoup
from re import findall

import datetime as dt
import openpyxl
import asyncio
import logging
import os


class Parser:
    def __init__(self, filename: str):
        self.workbook = openpyxl.load_workbook(f'./bot/uploads/{filename}')
        self.date = dt.datetime.strptime(f'{filename.split('.xlsx')[0]}{dt.date.today().year}', "%d.%m%Y").date()
        self.weekday = self.date.weekday()
        self.regular_lessons = []
        self.uday_lessons = []
        os.remove(f'./bot/uploads/{filename}')

    def parent_of_merged_cell(self, cell: MergedCell) -> str:
        """
        Ищет родителя объединённой клетки таблицы.

        Аргументы:
            cell (MergedCell): Объединённая клетка.

        Возвращает:
            str: Координаты родительской клетки.
        """
        sheet = cell.parent
        child_coord = cell.coordinate
        for merged in sheet.merged_cells.ranges:
            if child_coord in merged:
                return merged.start_cell.coordinate

    def cell_value(self, cell: Union[Cell, MergedCell]) -> str:
        """
        Возвращает значение клетки.

        Функция проверяет тип клетки, возвращает её значение если это обычная клетка, ищет родительскую клетку
        и возвращает её значение если клетка совмещённая.

        Аргументы:
            cell Union[Cell, MergedCell]: Клетка с искомым значением.

        Возвращает:
            str: Значение клетки.
        """
        if isinstance(cell, openpyxl.cell.cell.Cell):
            return cell.value
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            coord = self.parent_of_merged_cell(cell)
            parent = cell.parent[coord]
            return parent.value

    def regular_classes_schedule_parsing(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                         times_list: list, start_row: int, start_col: int, end_row: int,
                                         end_col: int) -> None:
        """
        Парсит обычное расписание классов.

        Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

        Аргументы:
            date (date): Дата из файла с расписанием.
            worksheet (Worksheet): Страница эксель файла.
            times_list (list): Список с таймингами уроков.
            start_row (int): Начальная строка итерации.
            start_col (int): Начальный столбец итерации.
            end_row (int): Конечная строка итерации.
            end_col (int): Конечный столбец итерации.

        Возвращает:
            None: функция ничего не возвращает.
        """
        for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            class_letter = self.cell_value(col[0])
            groups = ('гр.А', 'гр.Б')
            class_group = groups.index(''.join(self.cell_value(col[1]).split()))
            for num, cell in enumerate(col[2:]):
                if self.cell_value(cell):
                    lesson_info = '\n'.join((times_list[num], '\n'.join(self.cell_value(cell).split('\n\n'))))
                    self.regular_lessons.append(Lesson(num=num, info=lesson_info, date=self.date, group_num=class_group,
                                                class_letter=class_letter))

    def uday_groups_schedule_parsing(self, worksheet: Worksheet, times_list: list, start_row: int,
                                     start_col: int, end_row: int, end_col: int) -> None:
        """
        Парсит расписание для групп на универдень.

        Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

        Аргументы:
            worksheet (Worksheet): Страница эксель файла.
            times_list (list): Список с таймингами уроков.
            start_row (int): Начальная строка итерации.
            start_col (int): Начальный столбец итерации.
            end_row (int): Конечная строка итерации.
            end_col (int): Конечный столбец итерации.

        Возвращает:
            None: функция ничего не возвращает.
        """
        processed = set()
        for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            group_num = int(self.cell_value(col[0]).split()[0])
            if group_num in processed:
                continue
            processed.add(group_num)
            for num, cell in enumerate(col[1:]):
                if self.cell_value(cell):
                    lesson_info = '\n'.join((times_list[num], '\n'.join(self.cell_value(cell).split('\n\n'))))
                    self.uday_lessons.append(Lesson(num=num, info=lesson_info, date=self.date, group_num=group_num))

    def uday_classes_schedule_parsing(self, worksheet: Worksheet, times_list: list, start_row: int,
                                      start_col: int, end_row: int, end_col: int) -> None:
        """
        Парсит расписание для классов на универдень.

        Функция итерируется по столбцам и клеткам столбца, читает расписание и сохраняет его в базу данных.

        Аргументы:
            date (date): Дата из файла с расписанием.
            worksheet (Worksheet): Страница эксель файла.
            times_list (list): Список с таймингами уроков.
            start_row (int): Начальная строка итерации.
            start_col (int): Начальный столбец итерации.
            end_row (int): Конечная строка итерации.
            end_col (int): Конечный столбец итерации.

        Возвращает:
            None: функция ничего не возвращает.
        """
        processed = set()
        for col in worksheet.iter_cols(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            class_letter = self.cell_value(col[0])
            if class_letter in processed:
                continue
            processed.add(class_letter)
            for num, cell in enumerate(col[1:]):
                if self.cell_value(cell):
                    lesson_info = '\n'.join((times_list[num], '\n'.join(self.cell_value(cell).split('\n\n'))))
                    self.regular_lessons.append(Lesson(num=num, info=lesson_info, date=self.date, group_num=0,
                                                class_letter=class_letter))

    def parse(self) -> str:
        """
        Парсит .xlsx файл с расписанием.

        Удаляет при наличии устаревшие расписания из базы данных а также неактуальную версию расписания
        при его повторной загрузке, вызывает функции для парсинга различных листов файла, сохраняет
        информацию об уроках в атрибутах класса и вызывает функцию сохранения уроков в бд.

        Аргументы:
            None: Метод не принимает аргументов.

        Возвращает:
            str: Результат выполнения метода, сообщение об успехе или ошибке.
        """
        # Удаляем устаревшие расписаний
        old_date = dt.date.today() - dt.timedelta(days=2)
        delete_old_schedules(old_date)

        # Удаляем расписания при повторной загрузке расписания
        delete_repeated_schedule(self.date)

        # Определяем на какой странице чьё расписание
        sh_10, sh_11 = None, None
        for i, sh in enumerate(self.workbook.sheetnames):
            if sh.strip() == '10':
                sh_10 = i
            elif sh.strip() == '11':
                sh_11 = i
            if sh_10 and sh_11:
                break

        # Парсинг расписания 10-классников:
        try:
            # Универ-день (понедельник)
            if self.weekday == 0:
                sheet = self.workbook.worksheets[0] if not sh_10 else self.workbook.worksheets[sh_10]
                times_10 = [self.cell_value(i).replace('\n', '') for i in sheet['c'][2:9] + sheet['c'][9:11]
                            if self.cell_value(i)]
                self.uday_groups_schedule_parsing(sheet, times_10[:6], 2, 4, 8, 27)
                self.uday_classes_schedule_parsing(sheet, times_10[6:], 9, 4, 12, 27)

            # Все остальные дни недели
            else:
                sheet = self.workbook.worksheets[1] if not isinstance(sh_10, int) else self.workbook.worksheets[sh_10]
                ls_num = max([int(self.cell_value(i)) for i in sheet['B'][3:14] if str(self.cell_value(i)).isnumeric()])
                times_10 = [self.cell_value(i).replace('\n', '') for i in sheet['c'][3:3 + ls_num]]
                self.regular_classes_schedule_parsing(sheet, times_10, 2, 4, 11, 23)

        except Exception as ex:
            logging.error(ex)
            return 'Ошибка при парсинге расписания 10-х классов!'

        # Парсинг расписания 11-классников:
        try:
            # Универ-день (среда)
            if self.weekday == 2:
                sheet = self.workbook.worksheets[0] if not sh_11 else self.workbook.worksheets[sh_11]
                times_11 = [self.cell_value(i).replace('\n', '') for i in sheet['c'][3:9] + sheet['c'][10:12]
                            if self.cell_value(i)]
                self.uday_groups_schedule_parsing(sheet, times_11[:6], 3, 4, 9, 23)
                self.uday_classes_schedule_parsing(sheet, times_11[6:], 10, 4, 13, 23)

            # Все остальные дни недели
            else:
                sheet = self.workbook.worksheets[1] if not isinstance(sh_11, int) else self.workbook.worksheets[sh_11]
                times_11 = [self.cell_value(i).replace('\n', '') for i in sheet['c'][3:11] if self.cell_value(i)]
                self.regular_classes_schedule_parsing(sheet, times_11, 2, 4, 12, 23)

        except Exception as ex:
            logging.error(ex)
            return 'Ошибка при парсинге расписания 11-х классов!'

        return self.saving_to_database()

    def saving_to_database(self) -> str:
        """
        Сохраняет уроки полученные в ходе парсинга в базу данных.

        Аргументы:
            None: Метод не принимает аргументов.

        Возвращает:
            str: Сообщение о результате сохранения расписания в базу данных.
        """
        try:
            add_regular_lessons(self.regular_lessons)
            if self.uday_lessons:
                add_uday_lessons(self.uday_lessons)
        except Exception as ex:
            return f'Ошибка при сохранении расписания в базу данных!\nОшибка:\n{ex}'

        return 'Расписание сохранено успешно!'
    

async def parse_schedule_from_eljur(today, tomorrow, bot: Bot):
    env_vars = dotenv_values(".env")
    eljur_login, eljur_password = env_vars['ELJUR_LOGIN'], env_vars['ELJUR_PASSWORD']
    session = Session()
    url1 = f"https://fms.eljur.ru/ajaxauthorize"
    url2 = f"https://fms.eljur.ru/journal-board-action"
    session.post(url=url1, data={'username': eljur_login, 'password': eljur_password})
    page = session.get(url=url2)
    soup = BeautifulSoup(page.text, "html.parser")

    for obj in soup.find_all('div', class_='board-item__content'):
        text = obj.p.get_text()
        if 'расписание' in text.lower():
            date = findall(r'[0-9]{2}.[0-9]{2}', text)[0]
            if date == tomorrow.strftime('%d.%m'):
                schedule_file_url = obj.a['href']
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
                schedule = session.get(url=schedule_file_url, headers=headers)
                with open(f'./bot/uploads/{date}.xlsx', 'wb') as res:
                    res.write(schedule.content)
                parsing_result = Parser(f'{date}.xlsx').parse()                    
                if parsing_result == 'Расписание сохранено успешно!':
                    students = await get_all_users()
                    for student_id in students:
                        try:
                            await bot.send_message(chat_id=student_id, text=f'загружено расписание на завтра🗓')
                        except TelegramForbiddenError:
                            await delete_user(student_id)

                        # Задержка для избежения нарушения ограничений телеграма
                        await asyncio.sleep(0.035)
                    return True

            elif dt.datetime.strptime(date, '%d/%m') < today:
                break

    return False
